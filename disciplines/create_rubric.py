#!/usr/bin/env python3
"""
AI Workforce Automation Assessment Rubric Generator

Creates a rigorous, repeatable assessment framework for evaluating
job function automation potential using openpyxl.

Author: Kagami Analysis Framework
Date: January 2026
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment,
    NamedStyle, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime

# Assessment data based on rigorous analysis
DISCIPLINES = [
    {
        'code': 'QA',
        'name': 'Quality Assurance Engineer',
        'automation_pct': 85,
        'fully_automated': [
            'Load testing (k6/Gatling)',
            'Chaos experiments (230+ fault types)',
            'AI test generation',
            'Flaky test detection',
            'Coverage analysis',
            'Mutation testing',
            'Safety verification'
        ],
        'partially_automated': [
            'Exploratory test planning',
            'Production observability'
        ],
        'human_only': [
            'Test strategy definition',
            'Final release approval'
        ],
        'category': 'Quality'
    },
    {
        'code': 'ENG',
        'name': 'Software Engineer',
        'automation_pct': 80,
        'fully_automated': [
            'PR review (CodeAnt AI)',
            'Security scanning (Snyk, Endor Labs)',
            'Performance profiling',
            'Tech debt identification',
            'Refactoring planning',
            'Code generation'
        ],
        'partially_automated': [
            'ADR generation',
            'Architecture review',
            'Expert PR review'
        ],
        'human_only': [
            'Final architecture decisions',
            'Production deployment approval',
            'Complex system design'
        ],
        'category': 'Technical'
    },
    {
        'code': 'TPM',
        'name': 'Technical Program Manager',
        'automation_pct': 75,
        'fully_automated': [
            'Dependency graph building',
            'Critical path calculation',
            'Monte Carlo simulation',
            'Status report generation',
            'SLA-based escalation checks'
        ],
        'partially_automated': [
            'Risk estimation',
            'AI risk forecast'
        ],
        'human_only': [
            'Final escalation decisions',
            'Stakeholder communication',
            'Strategic prioritization'
        ],
        'category': 'Planning'
    },
    {
        'code': 'DE',
        'name': 'Data Engineer',
        'automation_pct': 75,
        'fully_automated': [
            'Pipeline generation (Airflow/Prefect)',
            'Data quality monitoring',
            'Schema registration',
            'Catalog management',
            'Query optimization'
        ],
        'partially_automated': [
            'CDC setup',
            'Cost analysis'
        ],
        'human_only': [
            'Data platform architecture',
            'Compliance validation',
            'Data governance policies'
        ],
        'category': 'Data'
    },
    {
        'code': 'PM',
        'name': 'Product Manager',
        'automation_pct': 70,
        'fully_automated': [
            'RICE+EFE prioritization scoring',
            'OKR tracking',
            'Epic decomposition',
            'Linear issue creation'
        ],
        'partially_automated': [
            'Roadmap generation',
            'PRD generation',
            'Customer signal synthesis',
            'Competitive analysis'
        ],
        'human_only': [
            'Strategic vision setting',
            'Customer interviews',
            'Final roadmap approval',
            'Stakeholder alignment'
        ],
        'category': 'Strategy'
    },
    {
        'code': 'DS',
        'name': 'Data Scientist',
        'automation_pct': 70,
        'fully_automated': [
            'Hyperparameter optimization',
            'Feature management',
            'Model registration',
            'Drift detection'
        ],
        'partially_automated': [
            'Experiment execution',
            'Causal inference',
            'Fairness audit'
        ],
        'human_only': [
            'Research question formulation',
            'Methodology selection',
            'Final model approval',
            'Business interpretation'
        ],
        'category': 'Research'
    },
    {
        'code': 'Design',
        'name': 'Product Designer',
        'automation_pct': 70,
        'fully_automated': [
            'Design system enforcement',
            'Accessibility auditing (WCAG)',
            'Component generation',
            'Motion specification'
        ],
        'partially_automated': [
            'Iterative design loops',
            'AI critique loop'
        ],
        'human_only': [
            'Initial design vision',
            'User research',
            'Final design approval',
            'Design strategy',
            'Aesthetic judgment'
        ],
        'category': 'Creative'
    },
    {
        'code': 'PMM',
        'name': 'Product Marketing Manager',
        'automation_pct': 65,
        'fully_automated': [
            'Content generation',
            'Messaging test design',
            'Competitive monitoring'
        ],
        'partially_automated': [
            'Positioning generation',
            'Competitive analysis',
            'Launch planning'
        ],
        'human_only': [
            'Brand strategy',
            'Market research',
            'Customer interviews',
            'Final messaging approval',
            'Go-to-market strategy'
        ],
        'category': 'Marketing'
    },
    {
        'code': 'EM',
        'name': 'Engineering Manager',
        'automation_pct': 60,
        'ceiling': 75,
        'fully_automated': [
            'Capacity calculation',
            'Team health metrics aggregation',
            'Performance metrics tracking'
        ],
        'partially_automated': [
            '1:1 prep (needs conversation)',
            'Allocation optimization',
            'Burnout prediction'
        ],
        'human_only': [
            'ALL people decisions',
            'Hiring decisions',
            'Firing decisions',
            'Promotions',
            'Performance ratings',
            'Difficult conversations',
            'Career development planning',
            'Culture building'
        ],
        'category': 'People'
    },
    {
        'code': 'UXR',
        'name': 'UX Researcher',
        'automation_pct': 55,
        'fully_automated': [
            'Interview transcription',
            'Survey design',
            'Insight synthesis',
            'Persona generation'
        ],
        'partially_automated': [
            'Usability testing (structure)',
            'Interview synthesis'
        ],
        'human_only': [
            'Research planning',
            'Participant recruitment',
            'Interview facilitation',
            'Final insight interpretation',
            'Research strategy',
            'Building rapport',
            'Reading non-verbal cues'
        ],
        'category': 'Research'
    }
]

def create_rubric():
    wb = Workbook()
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    
    subheader_font = Font(name='Calibri', size=11, bold=True)
    subheader_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    
    data_font = Font(name='Calibri', size=10)
    
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # ===== SHEET 1: Summary =====
    ws = wb.active
    ws.title = 'Summary'
    
    # Title
    ws['A1'] = 'AI Workforce Automation Assessment'
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='2F5496')
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    
    # Headers
    headers = ['Code', 'Role', 'Category', 'Automation %', 'Ceiling', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Data
    for row, d in enumerate(DISCIPLINES, 5):
        ws.cell(row=row, column=1, value=d['code']).border = thin_border
        ws.cell(row=row, column=2, value=d['name']).border = thin_border
        ws.cell(row=row, column=3, value=d['category']).border = thin_border
        
        pct_cell = ws.cell(row=row, column=4, value=d['automation_pct'] / 100)
        pct_cell.number_format = '0%'
        pct_cell.border = thin_border
        pct_cell.alignment = center_align
        
        ceiling = d.get('ceiling')
        ceiling_cell = ws.cell(row=row, column=5, value=ceiling / 100 if ceiling else None)
        if ceiling:
            ceiling_cell.number_format = '0%'
        ceiling_cell.border = thin_border
        ceiling_cell.alignment = center_align
        
        # Status based on automation level
        if d['automation_pct'] >= 75:
            status = 'High'
            status_fill = green_fill
        elif d['automation_pct'] >= 65:
            status = 'Medium'
            status_fill = yellow_fill
        else:
            status = 'Low'
            status_fill = red_fill
        
        status_cell = ws.cell(row=row, column=6, value=status)
        status_cell.fill = status_fill
        status_cell.border = thin_border
        status_cell.alignment = center_align
    
    # Average row
    avg_row = len(DISCIPLINES) + 5
    ws.cell(row=avg_row, column=1, value='').border = thin_border
    ws.cell(row=avg_row, column=2, value='AVERAGE').font = subheader_font
    ws.cell(row=avg_row, column=2).border = thin_border
    ws.cell(row=avg_row, column=3, value='').border = thin_border
    
    avg_formula = f'=AVERAGE(D5:D{avg_row-1})'
    avg_cell = ws.cell(row=avg_row, column=4, value=avg_formula)
    avg_cell.number_format = '0%'
    avg_cell.font = subheader_font
    avg_cell.border = thin_border
    avg_cell.alignment = center_align
    
    ws.cell(row=avg_row, column=5, value='').border = thin_border
    ws.cell(row=avg_row, column=6, value='').border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    
    # Add chart
    chart = BarChart()
    chart.type = 'bar'
    chart.style = 10
    chart.title = 'Automation by Role'
    chart.y_axis.title = 'Role'
    chart.x_axis.title = 'Automation %'
    
    data = Reference(ws, min_col=4, min_row=4, max_row=avg_row-1, max_col=4)
    cats = Reference(ws, min_col=1, min_row=5, max_row=avg_row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 12
    chart.height = 8
    
    ws.add_chart(chart, 'H4')
    
    # ===== SHEET 2: Detailed Breakdown =====
    ws2 = wb.create_sheet('Detailed Breakdown')
    
    ws2['A1'] = 'Task-Level Automation Breakdown'
    ws2['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    ws2.merge_cells('A1:E1')
    
    row = 3
    for d in DISCIPLINES:
        # Role header
        ws2.cell(row=row, column=1, value=f"{d['code']} - {d['name']}")
        ws2.cell(row=row, column=1).font = Font(name='Calibri', size=12, bold=True)
        ws2.cell(row=row, column=5, value=f"{d['automation_pct']}%")
        ws2.cell(row=row, column=5).font = Font(name='Calibri', size=12, bold=True)
        ws2.cell(row=row, column=5).alignment = center_align
        row += 1
        
        # Fully automated
        ws2.cell(row=row, column=1, value='Fully Automated')
        ws2.cell(row=row, column=1).font = subheader_font
        ws2.cell(row=row, column=1).fill = green_fill
        row += 1
        for task in d['fully_automated']:
            ws2.cell(row=row, column=2, value=f'• {task}')
            row += 1
        
        # Partially automated
        ws2.cell(row=row, column=1, value='Partially Automated')
        ws2.cell(row=row, column=1).font = subheader_font
        ws2.cell(row=row, column=1).fill = yellow_fill
        row += 1
        for task in d['partially_automated']:
            ws2.cell(row=row, column=2, value=f'• {task}')
            row += 1
        
        # Human only
        ws2.cell(row=row, column=1, value='Human Required')
        ws2.cell(row=row, column=1).font = subheader_font
        ws2.cell(row=row, column=1).fill = red_fill
        row += 1
        for task in d['human_only']:
            ws2.cell(row=row, column=2, value=f'• {task}')
            row += 1
        
        # Spacing
        row += 1
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['E'].width = 12
    
    # ===== SHEET 3: Scoring Rubric =====
    ws3 = wb.create_sheet('Scoring Rubric')
    
    ws3['A1'] = 'Automation Assessment Scoring Rubric'
    ws3['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    ws3.merge_cells('A1:D1')
    
    ws3['A3'] = 'Instructions: For each task, assign a score based on the criteria below'
    ws3['A3'].font = Font(name='Calibri', size=10, italic=True)
    
    # Rubric headers
    rubric_headers = ['Score', 'Classification', 'Criteria', 'Weight']
    for col, header in enumerate(rubric_headers, 1):
        cell = ws3.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Rubric data
    rubric_data = [
        (3, 'Fully Automated', 'Task can be performed entirely by AI without human intervention', 1.0),
        (2, 'Partially Automated', 'AI generates output but requires human review/approval', 0.66),
        (1, 'AI-Assisted', 'AI provides recommendations but human performs the task', 0.33),
        (0, 'Human Only', 'Task fundamentally requires human judgment, ethics, or relationship', 0.0),
    ]
    
    fills = [green_fill, yellow_fill, PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid'), red_fill]
    
    for i, (score, classification, criteria, weight) in enumerate(rubric_data):
        row = 6 + i
        ws3.cell(row=row, column=1, value=score).border = thin_border
        ws3.cell(row=row, column=1).alignment = center_align
        ws3.cell(row=row, column=1).fill = fills[i]
        
        ws3.cell(row=row, column=2, value=classification).border = thin_border
        ws3.cell(row=row, column=2).fill = fills[i]
        
        ws3.cell(row=row, column=3, value=criteria).border = thin_border
        ws3.cell(row=row, column=3).alignment = wrap_align
        
        ws3.cell(row=row, column=4, value=weight).border = thin_border
        ws3.cell(row=row, column=4).number_format = '0%'
        ws3.cell(row=row, column=4).alignment = center_align
    
    # Formula explanation
    ws3['A12'] = 'Automation % Calculation:'
    ws3['A12'].font = subheader_font
    
    ws3['A13'] = '= (Sum of Task Weights × Task Scores) / (Total Tasks × 3) × 100'
    ws3['A13'].font = Font(name='Consolas', size=10)
    
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 60
    ws3.column_dimensions['D'].width = 10
    
    # ===== SHEET 4: Assessment Template =====
    ws4 = wb.create_sheet('Assessment Template')
    
    ws4['A1'] = 'New Role Automation Assessment'
    ws4['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    
    ws4['A3'] = 'Role Name:'
    ws4['A4'] = 'Category:'
    ws4['A5'] = 'Assessor:'
    ws4['A6'] = 'Date:'
    
    ws4['B3'].border = thin_border
    ws4['B4'].border = thin_border
    ws4['B5'].border = thin_border
    ws4['B6'].border = thin_border
    
    # Task list
    headers = ['Task', 'Score (0-3)', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Empty task rows
    for row in range(9, 30):
        for col in range(1, 4):
            ws4.cell(row=row, column=col).border = thin_border
    
    # Total formula
    ws4['A31'] = 'Total Score:'
    ws4['A31'].font = subheader_font
    ws4['B31'] = '=SUM(B9:B29)'
    ws4['B31'].border = thin_border
    
    ws4['A32'] = 'Max Score:'
    ws4['A32'].font = subheader_font
    ws4['B32'] = '=COUNTA(A9:A29)*3'
    ws4['B32'].border = thin_border
    
    ws4['A33'] = 'Automation %:'
    ws4['A33'].font = Font(name='Calibri', size=12, bold=True)
    ws4['B33'] = '=IF(B32>0,B31/B32,0)'
    ws4['B33'].number_format = '0%'
    ws4['B33'].font = Font(name='Calibri', size=12, bold=True)
    ws4['B33'].border = thin_border
    
    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 40
    
    # Save
    output_path = '/Users/schizodactyl/projects/art/disciplines/automation_assessment_rubric.xlsx'
    wb.save(output_path)
    print(f'Rubric saved to: {output_path}')
    return output_path

if __name__ == '__main__':
    create_rubric()
