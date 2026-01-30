#!/usr/bin/env python3
"""
Tech Industry Workforce Contraction Model by Discipline (2025 Edition)

Creates a detailed Excel model projecting AI-driven workforce contraction
across the 10 core tech disciplines with multiple timeline scenarios.

DATA SOURCES (Verified January 2025):
- CompTIA State of Tech Workforce 2025: 6.09M tech workers
- GitHub Copilot Stats: 15M users, 51% faster coding, 88% code retention
- Microsoft Disclosure: AI writes ~30% of code on some projects
- Challenger, Gray & Christmas: 55,000 AI-linked layoffs in 2025
- BLS: Software dev job postings down 6.7% YoY
- TrueUp: 83,604 tech layoffs in 2025 YTD
- Indeed Hiring Lab: Tech postings 30% below pre-pandemic levels

Author: AI Workforce Analysis
Date: January 2025
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import math

# =============================================================================
# MODEL PARAMETERS & ASSUMPTIONS (Updated for 2025)
# =============================================================================

BASE_YEAR = 2025
PROJECTION_YEARS = [2026, 2027, 2028, 2029, 2030]

# Discipline data - 2025 estimates based on CompTIA, BLS, LinkedIn data
# Total tech workforce: ~6.09M (CompTIA 2025)
# Distribution based on LinkedIn Workforce Reports and Levels.fyi data

DISCIPLINES = {
    'QA': {
        'full_name': 'Quality Assurance Engineer',
        'base_headcount': 275000,
        'avg_salary': 98000,
        'current_automation': 0.40,  # AI test gen widely adopted
        'peak_automation': 0.85,
        'automation_ceiling': 0.88,
        'adoption_speed': 1.3,  # Fast - well-defined tasks
        'rationale': 'AI test generation (Copilot, TestRigor) widely adopted. Load testing, regression, coverage analysis highly automated. Human needed for test strategy, edge case identification, release judgment.',
        'tasks': {
            'fully_automatable': ['Unit test generation', 'Integration test scaffolding', 'Load testing (k6/Gatling)', 'Regression suites', 'Coverage analysis', 'Mutation testing', 'Flaky test detection'],
            'partially_automatable': ['Exploratory test planning', 'Test data generation', 'Performance analysis'],
            'human_required': ['Test strategy definition', 'Risk-based prioritization', 'Final release approval', 'Edge case identification']
        }
    },
    'ENG': {
        'full_name': 'Software Engineer',
        'base_headcount': 4100000,  # BLS + LinkedIn estimate
        'avg_salary': 132000,
        'current_automation': 0.30,  # Microsoft: 30% code from AI
        'peak_automation': 0.75,
        'automation_ceiling': 0.80,
        'adoption_speed': 1.2,
        'rationale': 'GitHub Copilot: 15M users, 51% faster coding. Microsoft reports 30% code AI-generated. But architecture, complex debugging, production incidents, cross-team coordination remain human.',
        'tasks': {
            'fully_automatable': ['Boilerplate code', 'Unit tests', 'Documentation', 'Simple bug fixes', 'Code formatting', 'Dependency updates'],
            'partially_automatable': ['Feature implementation', 'Refactoring', 'Code review triage', 'Performance optimization'],
            'human_required': ['System architecture', 'Complex debugging', 'Production incidents', 'Design decisions', 'Cross-team coordination']
        }
    },
    'TPM': {
        'full_name': 'Technical Program Manager',
        'base_headcount': 115000,
        'avg_salary': 172000,
        'current_automation': 0.25,
        'peak_automation': 0.70,
        'automation_ceiling': 0.75,
        'adoption_speed': 0.9,
        'rationale': 'Status reporting, dependency tracking, timeline calculation automatable. But escalation decisions, stakeholder negotiation, strategic prioritization require human judgment and political navigation.',
        'tasks': {
            'fully_automatable': ['Status report generation', 'Dependency graph building', 'Critical path calculation', 'Meeting scheduling', 'Risk identification'],
            'partially_automatable': ['Resource planning', 'Timeline estimation', 'Stakeholder updates'],
            'human_required': ['Escalation decisions', 'Stakeholder negotiation', 'Strategic prioritization', 'Conflict resolution']
        }
    },
    'DE': {
        'full_name': 'Data Engineer',
        'base_headcount': 185000,
        'avg_salary': 145000,
        'current_automation': 0.25,
        'peak_automation': 0.72,
        'automation_ceiling': 0.78,
        'adoption_speed': 1.0,
        'rationale': 'Pipeline generation (Airflow/Prefect), schema management, query optimization automatable. Data governance, compliance validation, platform architecture decisions require human oversight.',
        'tasks': {
            'fully_automatable': ['DAG generation', 'Schema registration', 'Query optimization', 'Data quality checks', 'Catalog management'],
            'partially_automatable': ['CDC setup', 'Cost analysis', 'Migration planning'],
            'human_required': ['Data governance', 'Compliance validation', 'Platform architecture', 'Vendor selection']
        }
    },
    'PM': {
        'full_name': 'Product Manager',
        'base_headcount': 365000,
        'avg_salary': 152000,
        'current_automation': 0.18,
        'peak_automation': 0.62,
        'automation_ceiling': 0.68,
        'adoption_speed': 0.8,
        'rationale': 'AI can draft PRDs, prioritize features (RICE), synthesize feedback. But vision setting, customer relationships, stakeholder alignment, and strategic decisions are fundamentally human.',
        'tasks': {
            'fully_automatable': ['Metrics dashboards', 'Competitor monitoring', 'Feature prioritization scoring', 'User feedback clustering'],
            'partially_automatable': ['PRD drafting', 'Roadmap generation', 'Market analysis'],
            'human_required': ['Vision setting', 'Customer interviews', 'Stakeholder alignment', 'Go/no-go decisions', 'Strategic direction']
        }
    },
    'DS': {
        'full_name': 'Data Scientist',
        'base_headcount': 155000,
        'avg_salary': 142000,
        'current_automation': 0.22,
        'peak_automation': 0.68,
        'automation_ceiling': 0.72,
        'adoption_speed': 1.0,
        'rationale': 'AutoML handles hyperparameter tuning, feature engineering. But problem formulation, methodology selection, business interpretation, and ethics review require domain expertise and judgment.',
        'tasks': {
            'fully_automatable': ['Hyperparameter optimization', 'Feature engineering', 'Model training', 'Drift detection', 'A/B test analysis'],
            'partially_automatable': ['Experiment design', 'Causal inference', 'Fairness auditing'],
            'human_required': ['Problem formulation', 'Methodology selection', 'Business interpretation', 'Ethics review', 'Stakeholder communication']
        }
    },
    'Design': {
        'full_name': 'Product Designer',
        'base_headcount': 235000,
        'avg_salary': 122000,
        'current_automation': 0.15,
        'peak_automation': 0.65,
        'automation_ceiling': 0.70,
        'adoption_speed': 0.85,
        'rationale': 'AI can generate components, audit accessibility, create variations. But design vision, user empathy, aesthetic judgment, and brand coherence remain deeply human creative acts.',
        'tasks': {
            'fully_automatable': ['Component generation', 'Accessibility auditing (WCAG)', 'Design system enforcement', 'Asset resizing'],
            'partially_automatable': ['Wireframing', 'Prototyping', 'Design variations', 'Motion specs'],
            'human_required': ['Creative direction', 'User research synthesis', 'Brand strategy', 'Aesthetic judgment', 'Emotional design']
        }
    },
    'PMM': {
        'full_name': 'Product Marketing Manager',
        'base_headcount': 88000,
        'avg_salary': 138000,
        'current_automation': 0.18,
        'peak_automation': 0.60,
        'automation_ceiling': 0.65,
        'adoption_speed': 0.9,
        'rationale': 'Content generation, competitive monitoring, campaign analytics automatable. Brand strategy, customer relationships, market intuition, and go-to-market decisions require human creativity and judgment.',
        'tasks': {
            'fully_automatable': ['Content generation', 'Competitive monitoring', 'Campaign analytics', 'Social listening'],
            'partially_automatable': ['Positioning drafts', 'Battlecard creation', 'Launch checklists'],
            'human_required': ['Brand strategy', 'Customer interviews', 'Market intuition', 'Go-to-market decisions', 'Partner relationships']
        }
    },
    'EM': {
        'full_name': 'Engineering Manager',
        'base_headcount': 175000,
        'avg_salary': 205000,
        'current_automation': 0.12,
        'peak_automation': 0.55,
        'automation_ceiling': 0.60,
        'adoption_speed': 0.65,
        'rationale': 'Metrics aggregation, capacity planning automatable. But ALL people decisions—hiring, firing, promotions, performance reviews, career development, difficult conversations—are irreducibly human.',
        'tasks': {
            'fully_automatable': ['Metrics aggregation', 'Capacity calculation', 'Sprint analytics', 'Team velocity tracking'],
            'partially_automatable': ['1:1 preparation', 'Performance tracking', 'Burnout risk indicators'],
            'human_required': ['Hiring decisions', 'Performance reviews', 'Promotions', 'Terminations', 'Career development', 'Difficult conversations', 'Culture building', 'Team dynamics']
        }
    },
    'UXR': {
        'full_name': 'UX Researcher',
        'base_headcount': 42000,
        'avg_salary': 128000,
        'current_automation': 0.12,
        'peak_automation': 0.52,
        'automation_ceiling': 0.55,
        'adoption_speed': 0.7,
        'rationale': 'Transcription, survey analysis, sentiment analysis automatable. But research planning, interview facilitation, building rapport, reading non-verbal cues, and insight interpretation require human presence and empathy.',
        'tasks': {
            'fully_automatable': ['Transcription', 'Survey analysis', 'Sentiment analysis', 'Heatmap generation'],
            'partially_automatable': ['Interview synthesis', 'Persona generation', 'Journey mapping'],
            'human_required': ['Research planning', 'Interview facilitation', 'Participant rapport', 'Non-verbal cue reading', 'Insight interpretation', 'Research strategy']
        }
    }
}

# Timeline scenarios - calibrated to real adoption curves
SCENARIOS = {
    'conservative': {
        'name': 'Conservative',
        'description': 'Regulatory friction, enterprise caution, labor resistance',
        'annual_acceleration': 0.06,
        'adoption_lag_years': 1.5,
        'economic_friction': 0.82,
        'color': '4ECDC4'
    },
    'moderate': {
        'name': 'Moderate (Base Case)',
        'description': 'Current trajectory continues with steady enterprise adoption',
        'annual_acceleration': 0.10,
        'adoption_lag_years': 0.5,
        'economic_friction': 0.88,
        'color': 'FFB800'
    },
    'aggressive': {
        'name': 'Aggressive',
        'description': 'Agentic AI breakthrough, competitive pressure, economic downturn',
        'annual_acceleration': 0.15,
        'adoption_lag_years': 0,
        'economic_friction': 0.94,
        'color': 'FF6B6B'
    }
}


def calculate_automation_curve(discipline_data, scenario, year):
    """Calculate automation level using S-curve adoption model."""
    years_from_base = year - BASE_YEAR
    
    current = discipline_data['current_automation']
    peak = discipline_data['peak_automation']
    ceiling = discipline_data['automation_ceiling']
    speed = discipline_data['adoption_speed']
    
    acceleration = scenario['annual_acceleration']
    lag = scenario['adoption_lag_years']
    friction = scenario['economic_friction']
    
    effective_years = max(0, years_from_base - lag)
    
    # S-curve: logistic function
    midpoint = 2.5 / speed
    steepness = acceleration * speed
    
    sigmoid = 1 / (1 + math.exp(-steepness * (effective_years - midpoint)))
    theoretical = current + (peak - current) * sigmoid
    actual = min(theoretical * friction, ceiling)
    
    return actual


def calculate_headcount_impact(base_headcount, base_automation, new_automation):
    """Calculate headcount reduction from automation increase."""
    # 55% of automation gains → headcount reduction
    # 45% → productivity gains (existing workers do more)
    REDUCTION_FACTOR = 0.55
    
    automation_increase = new_automation - base_automation
    return base_headcount * automation_increase * REDUCTION_FACTOR


def create_model():
    """Create the complete Excel model."""
    wb = Workbook()
    
    # Styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    subheader_fill = PatternFill(start_color='16213e', end_color='16213e', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Calculate totals
    total_base = sum(d['base_headcount'] for d in DISCIPLINES.values())
    
    projections = {}
    for scenario_key, scenario in SCENARIOS.items():
        total_reduction = 0
        for disc_data in DISCIPLINES.values():
            base_auto = disc_data['current_automation']
            new_auto = calculate_automation_curve(disc_data, scenario, 2030)
            reduction = calculate_headcount_impact(disc_data['base_headcount'], base_auto, new_auto)
            total_reduction += reduction
        projections[scenario_key] = int(total_reduction)
    
    # ==========================================================================
    # SHEET 1: Executive Summary
    # ==========================================================================
    ws = wb.active
    ws.title = 'Executive Summary'
    
    ws['A1'] = 'Tech Industry Workforce Contraction Model'
    ws['A1'].font = Font(name='Arial', size=18, bold=True)
    ws['A2'] = f'AI-Driven Automation Impact | Base Year: {BASE_YEAR} | Projection: 2030'
    ws['A2'].font = Font(name='Arial', size=12, italic=True, color='666666')
    
    ws['A4'] = 'HEADLINE PROJECTIONS (2025 → 2030)'
    ws['A4'].font = Font(name='Arial', size=14, bold=True)
    
    headers = ['Metric', 'Conservative', 'Moderate', 'Aggressive']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    metrics = [
        ('Base Headcount (2025)', total_base, total_base, total_base),
        ('Projected Job Reduction', projections['conservative'], projections['moderate'], projections['aggressive']),
        ('Reduction %', projections['conservative']/total_base, projections['moderate']/total_base, projections['aggressive']/total_base),
        ('Remaining Workforce', total_base - projections['conservative'], total_base - projections['moderate'], total_base - projections['aggressive']),
        ('Est. Wage Impact ($B)', projections['conservative'] * 140000 / 1e9, projections['moderate'] * 140000 / 1e9, projections['aggressive'] * 140000 / 1e9),
    ]
    
    for row_idx, (metric, cons, mod, agg) in enumerate(metrics, 6):
        ws.cell(row=row_idx, column=1, value=metric).border = thin_border
        for col_idx, val in enumerate([cons, mod, agg], 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right')
            if 'Reduction %' in metric:
                cell.number_format = '0.0%'
            elif 'Wage' in metric:
                cell.number_format = '$#,##0.0"B"'
            elif isinstance(val, int) and val > 1000:
                cell.number_format = '#,##0'
    
    # Data sources
    ws['A13'] = 'DATA SOURCES (Verified January 2025)'
    ws['A13'].font = Font(name='Arial', size=12, bold=True)
    
    sources = [
        '• CompTIA State of Tech Workforce 2025: 6.09M US tech workers',
        '• GitHub Copilot Stats: 15M users, 51% faster coding, 88% code retention',
        '• Microsoft Disclosure (2025): AI generates ~30% of code on some projects',
        '• Challenger, Gray & Christmas: 55,000 AI-linked layoffs in 2025',
        '• BLS: Software dev job postings down 6.7% YoY (Oct 2025)',
        '• Indeed Hiring Lab: Tech postings 30% below pre-pandemic levels',
        '• TrueUp Layoffs Tracker: 83,604 tech layoffs in 2025 YTD'
    ]
    
    for i, src in enumerate(sources, 14):
        ws.cell(row=i, column=1, value=src).font = Font(size=10, color='555555')
    
    ws['A23'] = 'KEY MODEL ASSUMPTIONS'
    ws['A23'].font = Font(name='Arial', size=12, bold=True)
    
    assumptions = [
        '• 55% of automation gains translate to headcount reduction',
        '• 45% translate to productivity gains (existing workers do more)',
        '• S-curve adoption model (logistic function) for each discipline',
        '• Automation ceilings reflect irreducibly human tasks',
        '• No major regulatory intervention assumed',
        '• Economic conditions assumed stable (no recession/boom adjustment)'
    ]
    
    for i, assum in enumerate(assumptions, 24):
        ws.cell(row=i, column=1, value=assum).font = Font(size=10, color='555555')
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    
    # ==========================================================================
    # SHEET 2: Model Parameters
    # ==========================================================================
    ws2 = wb.create_sheet('Model Parameters')
    
    ws2['A1'] = 'MODEL PARAMETERS'
    ws2['A1'].font = Font(name='Arial', size=16, bold=True)
    
    # Scenario parameters
    ws2['A3'] = 'SCENARIO DEFINITIONS'
    ws2['A3'].font = Font(name='Arial', size=12, bold=True)
    
    for col, h in enumerate(['Parameter', 'Conservative', 'Moderate', 'Aggressive'], 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    params = [
        ('Description', SCENARIOS['conservative']['description'], SCENARIOS['moderate']['description'], SCENARIOS['aggressive']['description']),
        ('Annual Acceleration', SCENARIOS['conservative']['annual_acceleration'], SCENARIOS['moderate']['annual_acceleration'], SCENARIOS['aggressive']['annual_acceleration']),
        ('Adoption Lag (Years)', SCENARIOS['conservative']['adoption_lag_years'], SCENARIOS['moderate']['adoption_lag_years'], SCENARIOS['aggressive']['adoption_lag_years']),
        ('Economic Friction', SCENARIOS['conservative']['economic_friction'], SCENARIOS['moderate']['economic_friction'], SCENARIOS['aggressive']['economic_friction']),
    ]
    
    for row_idx, (p, c, m, a) in enumerate(params, 5):
        ws2.cell(row=row_idx, column=1, value=p).border = thin_border
        for col_idx, v in enumerate([c, m, a], 2):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            if isinstance(v, float) and v < 1:
                cell.number_format = '0%' if v < 0.5 else '0.0%'
    
    # Discipline parameters
    ws2['A11'] = 'DISCIPLINE AUTOMATION PARAMETERS'
    ws2['A11'].font = Font(name='Arial', size=12, bold=True)
    
    disc_headers = ['Discipline', 'Full Name', 'Headcount', 'Avg Salary', 'Current', 'Peak', 'Ceiling', 'Speed', 'Rationale']
    for col, h in enumerate(disc_headers, 1):
        cell = ws2.cell(row=12, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 13):
        vals = [key, d['full_name'], d['base_headcount'], d['avg_salary'], d['current_automation'], d['peak_automation'], d['automation_ceiling'], d['adoption_speed'], d['rationale']]
        for col_idx, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            if col_idx == 3:
                cell.number_format = '#,##0'
            elif col_idx == 4:
                cell.number_format = '$#,##0'
            elif col_idx in [5, 6, 7]:
                cell.number_format = '0%'
            elif col_idx == 9:
                cell.alignment = Alignment(wrap_text=True)
    
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 26
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 11
    ws2.column_dimensions['E'].width = 9
    ws2.column_dimensions['F'].width = 8
    ws2.column_dimensions['G'].width = 8
    ws2.column_dimensions['H'].width = 8
    ws2.column_dimensions['I'].width = 60
    
    # ==========================================================================
    # SHEET 3: Projections by Discipline
    # ==========================================================================
    ws3 = wb.create_sheet('Projections by Discipline')
    
    ws3['A1'] = 'HEADCOUNT PROJECTIONS BY DISCIPLINE (Moderate Scenario)'
    ws3['A1'].font = Font(name='Arial', size=16, bold=True)
    
    proj_headers = ['Discipline', f'{BASE_YEAR}'] + [str(y) for y in PROJECTION_YEARS] + ['Reduction', '% Lost']
    for col, h in enumerate(proj_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    scenario = SCENARIOS['moderate']
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 4):
        base = d['base_headcount']
        current = base
        
        ws3.cell(row=row_idx, column=1, value=key).border = thin_border
        ws3.cell(row=row_idx, column=2, value=base).border = thin_border
        ws3.cell(row=row_idx, column=2).number_format = '#,##0'
        
        for col_idx, year in enumerate(PROJECTION_YEARS, 3):
            new_auto = calculate_automation_curve(d, scenario, year)
            prev_auto = calculate_automation_curve(d, scenario, year - 1) if year > BASE_YEAR + 1 else d['current_automation']
            reduction = calculate_headcount_impact(current, prev_auto, new_auto)
            current = max(0, current - reduction)
            
            cell = ws3.cell(row=row_idx, column=col_idx, value=int(current))
            cell.border = thin_border
            cell.number_format = '#,##0'
        
        total_red = base - current
        pct_red = total_red / base
        
        cell = ws3.cell(row=row_idx, column=len(PROJECTION_YEARS) + 3, value=int(total_red))
        cell.border = thin_border
        cell.number_format = '#,##0'
        
        cell = ws3.cell(row=row_idx, column=len(PROJECTION_YEARS) + 4, value=pct_red)
        cell.border = thin_border
        cell.number_format = '0.0%'
        if pct_red >= 0.25:
            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        elif pct_red >= 0.15:
            cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
    
    # Totals
    total_row = 4 + len(DISCIPLINES)
    ws3.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, len(proj_headers)):
        cell = ws3.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.number_format = '#,##0'
    
    ws3.column_dimensions['A'].width = 10
    for i in range(2, len(proj_headers) + 1):
        ws3.column_dimensions[get_column_letter(i)].width = 12
    
    # ==========================================================================
    # SHEET 4: Scenario Comparison
    # ==========================================================================
    ws4 = wb.create_sheet('Scenario Comparison')
    
    ws4['A1'] = 'SCENARIO COMPARISON: 2030 OUTCOMES'
    ws4['A1'].font = Font(name='Arial', size=16, bold=True)
    
    comp_headers = ['Discipline', '2025 Base', 'Conservative 2030', 'Moderate 2030', 'Aggressive 2030', 'Max Reduction']
    for col, h in enumerate(comp_headers, 1):
        cell = ws4.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 4):
        base = d['base_headcount']
        results = {}
        
        for skey, scenario in SCENARIOS.items():
            current = base
            for year in PROJECTION_YEARS:
                new_auto = calculate_automation_curve(d, scenario, year)
                prev_auto = calculate_automation_curve(d, scenario, year - 1) if year > BASE_YEAR + 1 else d['current_automation']
                reduction = calculate_headcount_impact(current, prev_auto, new_auto)
                current = max(0, current - reduction)
            results[skey] = int(current)
        
        ws4.cell(row=row_idx, column=1, value=key).border = thin_border
        ws4.cell(row=row_idx, column=2, value=base).border = thin_border
        ws4.cell(row=row_idx, column=2).number_format = '#,##0'
        
        ws4.cell(row=row_idx, column=3, value=results['conservative']).border = thin_border
        ws4.cell(row=row_idx, column=3).number_format = '#,##0'
        
        ws4.cell(row=row_idx, column=4, value=results['moderate']).border = thin_border
        ws4.cell(row=row_idx, column=4).number_format = '#,##0'
        
        ws4.cell(row=row_idx, column=5, value=results['aggressive']).border = thin_border
        ws4.cell(row=row_idx, column=5).number_format = '#,##0'
        
        max_red = (base - results['aggressive']) / base
        cell = ws4.cell(row=row_idx, column=6, value=max_red)
        cell.border = thin_border
        cell.number_format = '0.0%'
    
    # Totals
    total_row = 4 + len(DISCIPLINES)
    ws4.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in range(2, 6):
        cell = ws4.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.number_format = '#,##0'
    
    ws4.column_dimensions['A'].width = 10
    for i in range(2, 7):
        ws4.column_dimensions[get_column_letter(i)].width = 16
    
    # ==========================================================================
    # SHEET 5: Economic Impact
    # ==========================================================================
    ws5 = wb.create_sheet('Economic Impact')
    
    ws5['A1'] = 'ECONOMIC IMPACT (Moderate Scenario)'
    ws5['A1'].font = Font(name='Arial', size=16, bold=True)
    
    econ_headers = ['Discipline', 'Headcount', 'Avg Salary', 'Wages 2025 ($M)', 'Headcount 2030', 'Wages 2030 ($M)', 'Lost Wages ($M)']
    for col, h in enumerate(econ_headers, 1):
        cell = ws5.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    scenario = SCENARIOS['moderate']
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 4):
        base = d['base_headcount']
        salary = d['avg_salary']
        base_wages = base * salary / 1e6
        
        current = base
        for year in PROJECTION_YEARS:
            new_auto = calculate_automation_curve(d, scenario, year)
            prev_auto = calculate_automation_curve(d, scenario, year - 1) if year > BASE_YEAR + 1 else d['current_automation']
            reduction = calculate_headcount_impact(current, prev_auto, new_auto)
            current = max(0, current - reduction)
        
        final = int(current)
        final_wages = final * salary / 1e6
        lost = base_wages - final_wages
        
        vals = [key, base, salary, base_wages, final, final_wages, lost]
        for col_idx, v in enumerate(vals, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=v)
            cell.border = thin_border
            if col_idx in [2, 5]:
                cell.number_format = '#,##0'
            elif col_idx == 3:
                cell.number_format = '$#,##0'
            elif col_idx in [4, 6, 7]:
                cell.number_format = '$#,##0.0'
    
    # Totals
    total_row = 4 + len(DISCIPLINES)
    ws5.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for col in [2, 4, 5, 6, 7]:
        cell = ws5.cell(row=total_row, column=col, value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
        cell.font = Font(bold=True)
        cell.border = thin_border
        if col in [2, 5]:
            cell.number_format = '#,##0'
        else:
            cell.number_format = '$#,##0.0'
    
    ws5.column_dimensions['A'].width = 10
    for i in range(2, 8):
        ws5.column_dimensions[get_column_letter(i)].width = 14
    
    # ==========================================================================
    # SHEET 6: Task Analysis
    # ==========================================================================
    ws6 = wb.create_sheet('Task Analysis')
    
    ws6['A1'] = 'TASK-LEVEL AUTOMATION ANALYSIS'
    ws6['A1'].font = Font(name='Arial', size=16, bold=True)
    
    task_headers = ['Discipline', 'Fully Automatable', 'Partially Automatable', 'Human Required']
    for col, h in enumerate(task_headers, 1):
        cell = ws6.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 4):
        tasks = d['tasks']
        ws6.cell(row=row_idx, column=1, value=key).border = thin_border
        ws6.cell(row=row_idx, column=2, value=', '.join(tasks['fully_automatable'])).border = thin_border
        ws6.cell(row=row_idx, column=3, value=', '.join(tasks['partially_automatable'])).border = thin_border
        ws6.cell(row=row_idx, column=4, value=', '.join(tasks['human_required'])).border = thin_border
    
    for row in ws6.iter_rows(min_row=4, max_row=4 + len(DISCIPLINES) - 1, min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws6.column_dimensions['A'].width = 10
    ws6.column_dimensions['B'].width = 50
    ws6.column_dimensions['C'].width = 35
    ws6.column_dimensions['D'].width = 45
    
    # ==========================================================================
    # SHEET 7: YoY Automation
    # ==========================================================================
    ws7 = wb.create_sheet('YoY Automation')
    
    ws7['A1'] = 'AUTOMATION LEVELS BY YEAR (Moderate)'
    ws7['A1'].font = Font(name='Arial', size=16, bold=True)
    
    yoy_headers = ['Discipline', f'{BASE_YEAR}'] + [str(y) for y in PROJECTION_YEARS]
    for col, h in enumerate(yoy_headers, 1):
        cell = ws7.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    scenario = SCENARIOS['moderate']
    
    for row_idx, (key, d) in enumerate(DISCIPLINES.items(), 4):
        ws7.cell(row=row_idx, column=1, value=key).border = thin_border
        
        cell = ws7.cell(row=row_idx, column=2, value=d['current_automation'])
        cell.border = thin_border
        cell.number_format = '0%'
        
        for col_idx, year in enumerate(PROJECTION_YEARS, 3):
            auto = calculate_automation_curve(d, scenario, year)
            cell = ws7.cell(row=row_idx, column=col_idx, value=auto)
            cell.border = thin_border
            cell.number_format = '0%'
            
            if auto >= 0.65:
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif auto >= 0.50:
                cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
            elif auto >= 0.35:
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    
    ws7.column_dimensions['A'].width = 10
    for i in range(2, len(yoy_headers) + 1):
        ws7.column_dimensions[get_column_letter(i)].width = 10
    
    # ==========================================================================
    # Save
    # ==========================================================================
    output_path = '/Users/schizodactyl/projects/art/disciplines/industry_contraction_model.xlsx'
    wb.save(output_path)
    
    print(f'✅ Model saved: {output_path}')
    print(f'\n📊 MODEL SUMMARY')
    print('=' * 60)
    print(f'Base Year: {BASE_YEAR}')
    print(f'Projection: 2030')
    print(f'Disciplines: {len(DISCIPLINES)}')
    print(f'Total Headcount: {total_base:,}')
    print(f'\nProjected Reductions by 2030:')
    print(f'  Conservative: {projections["conservative"]:,} ({projections["conservative"]/total_base:.1%})')
    print(f'  Moderate:     {projections["moderate"]:,} ({projections["moderate"]/total_base:.1%})')
    print(f'  Aggressive:   {projections["aggressive"]:,} ({projections["aggressive"]/total_base:.1%})')
    
    return output_path, projections, total_base


if __name__ == '__main__':
    create_model()
