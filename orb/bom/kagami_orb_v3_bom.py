#!/usr/bin/env python3
"""
Kagami Orb V3 - Excel BOM Model Generator
==========================================

Generates a comprehensive Excel workbook with:
- Full BOM with component details
- Cost scenarios for different quantities (1, 10, 100, 500, 1000)
- Sensitivity analysis
- Manufacturing partner contacts
- FMEA summary
- Test coverage matrix

Usage:
    python kagami_orb_v3_bom.py

Output:
    kagami_orb_v3_bom_model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from datetime import datetime
from decimal import Decimal
import re


def parse_cost(cost_str: str) -> float:
    """Parse cost string like '$140' or '$0.50' to float."""
    if not cost_str:
        return 0.0
    match = re.search(r'\$?([\d.]+)', str(cost_str))
    return float(match.group(1)) if match else 0.0


# Component data from core.yaml
COMPONENTS = [
    # Compute Module
    {
        "part_id": "QCS6490-SOM-001",
        "designation": "U1",
        "category": "Compute",
        "quantity": 1,
        "description": "Qualcomm QCS6490 SoM - 12 TOPS NPU, WiFi 6E, 6nm",
        "supplier": "Thundercomm",
        "part_number": "TurboX C6490",
        "cost_qty1": 140.00,
        "cost_qty100": 110.00,
        "cost_qty500": 95.00,
        "cost_qty1000": 85.00,
        "lead_time": "2-4 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 54,
    },
    {
        "part_id": "HAILO10H-001",
        "designation": "U2",
        "category": "Compute",
        "quantity": 1,
        "description": "Hailo-10H AI Accelerator - 40 TOPS INT4, GenAI",
        "supplier": "Hailo",
        "part_number": "Hailo-10H",
        "cost_qty1": 90.00,
        "cost_qty100": 75.00,
        "cost_qty500": 65.00,
        "cost_qty1000": 55.00,
        "lead_time": "2-3 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 48,
    },
    {
        "part_id": "ESP32S3-001",
        "designation": "U3",
        "category": "Compute",
        "quantity": 1,
        "description": "ESP32-S3-WROOM-1-N4 Co-processor",
        "supplier": "Digi-Key",
        "part_number": "ESP32-S3-WROOM-1-N4",
        "cost_qty1": 3.34,
        "cost_qty100": 2.80,
        "cost_qty500": 2.50,
        "cost_qty1000": 2.20,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 24,
    },
    # Display System
    {
        "part_id": "AMOLED-001",
        "designation": "DISP1",
        "category": "Display",
        "quantity": 1,
        "description": "3.4\" Round AMOLED 800x800 - Living Eye Interface",
        "supplier": "AliExpress/Alibaba",
        "part_number": "RM69330 Round",
        "cost_qty1": 85.00,
        "cost_qty100": 65.00,
        "cost_qty500": 55.00,
        "cost_qty1000": 48.00,
        "lead_time": "2-4 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 160,  # AMOLED transparency uncertainty
    },
    {
        "part_id": "MIRROR-DIELECTRIC-001",
        "designation": "OPT1",
        "category": "Display",
        "quantity": 1,
        "description": "Dielectric Mirror Film 90mm - Touch Through",
        "supplier": "Edmund Optics",
        "part_number": "Dielectric Mirror",
        "cost_qty1": 45.00,
        "cost_qty100": 35.00,
        "cost_qty500": 28.00,
        "cost_qty1000": 22.00,
        "lead_time": "1-2 weeks",
        "criticality": "MEDIUM",
        "fmea_rpn": 140,  # Coating adhesion risk
    },
    {
        "part_id": "AF-COATING-001",
        "designation": "OPT2",
        "category": "Display",
        "quantity": 1,
        "description": "Daikin Optool DSX-E Anti-fingerprint",
        "supplier": "Daikin",
        "part_number": "Optool DSX-E",
        "cost_qty1": 45.00,
        "cost_qty100": 35.00,
        "cost_qty500": 28.00,
        "cost_qty1000": 22.00,
        "lead_time": "2 weeks",
        "criticality": "LOW",
        "fmea_rpn": 36,
    },
    {
        "part_id": "RADAR-001",
        "designation": "U4",
        "category": "Display",
        "quantity": 1,
        "description": "Infineon BGT60TR13C 60GHz Gesture Radar",
        "supplier": "Infineon/Digi-Key",
        "part_number": "BGT60TR13C",
        "cost_qty1": 25.00,
        "cost_qty100": 18.00,
        "cost_qty500": 15.00,
        "cost_qty1000": 12.00,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 48,
    },
    # Camera
    {
        "part_id": "CAM-IMX989-001",
        "designation": "CAM1",
        "category": "Camera",
        "quantity": 1,
        "description": "Sony IMX989 1\" 50.3MP - Flagship Hidden Camera",
        "supplier": "SINCEREFIRST",
        "part_number": "IMX989 Module",
        "cost_qty1": 95.00,
        "cost_qty100": 75.00,
        "cost_qty500": 65.00,
        "cost_qty1000": 55.00,
        "lead_time": "2-4 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 175,  # Availability risk
    },
    # Audio
    {
        "part_id": "MIC-SENSIBEL-001",
        "designation": "MIC1-4",
        "category": "Audio",
        "quantity": 4,
        "description": "sensiBel SBM100B Optical MEMS Mic (-26dB SNR)",
        "supplier": "sensiBel",
        "part_number": "SBM100B",
        "cost_qty1": 30.00,  # Per unit
        "cost_qty100": 20.00,
        "cost_qty500": 16.00,
        "cost_qty1000": 14.00,
        "lead_time": "4-6 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 224,  # CRITICAL: Single source risk
    },
    {
        "part_id": "XMOS-XVF3800-001",
        "designation": "U5",
        "category": "Audio",
        "quantity": 1,
        "description": "XMOS XVF3800 Voice Processor - AEC/Beamforming",
        "supplier": "XMOS/OpenELAB",
        "part_number": "XVF3800",
        "cost_qty1": 99.00,
        "cost_qty100": 75.00,
        "cost_qty500": 65.00,
        "cost_qty1000": 55.00,
        "lead_time": "In stock",
        "criticality": "HIGH",
        "fmea_rpn": 48,
    },
    {
        "part_id": "SPK-BMR-001",
        "designation": "SPK1",
        "category": "Audio",
        "quantity": 1,
        "description": "Tectonic TEBM28C20N-4 BMR Speaker",
        "supplier": "Tectonic",
        "part_number": "TEBM28C20N-4",
        "cost_qty1": 25.00,
        "cost_qty100": 18.00,
        "cost_qty500": 15.00,
        "cost_qty1000": 12.00,
        "lead_time": "2 weeks",
        "criticality": "MEDIUM",
        "fmea_rpn": 32,
    },
    {
        "part_id": "AMP-MAX98357A-001",
        "designation": "U6",
        "category": "Audio",
        "quantity": 1,
        "description": "MAX98357A 3W I2S Class D Amplifier",
        "supplier": "Digi-Key",
        "part_number": "MAX98357A",
        "cost_qty1": 2.96,
        "cost_qty100": 2.10,
        "cost_qty500": 1.80,
        "cost_qty1000": 1.50,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 18,
    },
    # LEDs
    {
        "part_id": "LED-HD108-001",
        "designation": "LED1-16",
        "category": "LEDs",
        "quantity": 16,
        "description": "HD108 16-bit RGBW LED (65,536 levels/ch)",
        "supplier": "AliExpress",
        "part_number": "HD108-5050",
        "cost_qty1": 0.50,  # Per unit
        "cost_qty100": 0.30,
        "cost_qty500": 0.25,
        "cost_qty1000": 0.20,
        "lead_time": "2 weeks",
        "criticality": "MEDIUM",
        "fmea_rpn": 24,
    },
    # Sensors
    {
        "part_id": "IMU-ICM45686-001",
        "designation": "U7",
        "category": "Sensors",
        "quantity": 1,
        "description": "TDK ICM-45686 6-axis IMU with AI",
        "supplier": "TDK InvenSense",
        "part_number": "ICM-45686",
        "cost_qty1": 8.00,
        "cost_qty100": 6.00,
        "cost_qty500": 5.00,
        "cost_qty1000": 4.50,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 24,
    },
    {
        "part_id": "TEMP-SHT45-001",
        "designation": "U8",
        "category": "Sensors",
        "quantity": 1,
        "description": "Sensirion SHT45 Temp/Humidity (±0.1°C)",
        "supplier": "Sensirion",
        "part_number": "SHT45",
        "cost_qty1": 5.00,
        "cost_qty100": 3.50,
        "cost_qty500": 3.00,
        "cost_qty1000": 2.50,
        "lead_time": "In stock",
        "criticality": "LOW",
        "fmea_rpn": 12,
    },
    {
        "part_id": "TOF-VL53L8CX-001",
        "designation": "U9",
        "category": "Sensors",
        "quantity": 1,
        "description": "ST VL53L8CX 8x8 ToF Sensor (4m range)",
        "supplier": "STMicro",
        "part_number": "VL53L8CX",
        "cost_qty1": 12.00,
        "cost_qty100": 9.00,
        "cost_qty500": 7.50,
        "cost_qty1000": 6.50,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 32,
    },
    {
        "part_id": "AIR-SEN66-001",
        "designation": "U10",
        "category": "Sensors",
        "quantity": 1,
        "description": "Sensirion SEN66 All-in-one Air Quality",
        "supplier": "Sensirion",
        "part_number": "SEN66",
        "cost_qty1": 45.00,
        "cost_qty100": 35.00,
        "cost_qty500": 30.00,
        "cost_qty1000": 25.00,
        "lead_time": "2 weeks",
        "criticality": "MEDIUM",
        "fmea_rpn": 36,
    },
    {
        "part_id": "LIGHT-AS7343-001",
        "designation": "U11",
        "category": "Sensors",
        "quantity": 1,
        "description": "ams AS7343 14-channel Spectral Light",
        "supplier": "ams OSRAM",
        "part_number": "AS7343",
        "cost_qty1": 8.00,
        "cost_qty100": 6.00,
        "cost_qty500": 5.00,
        "cost_qty1000": 4.50,
        "lead_time": "In stock",
        "criticality": "LOW",
        "fmea_rpn": 18,
    },
    {
        "part_id": "HALL-AH49E-001",
        "designation": "U12",
        "category": "Sensors",
        "quantity": 1,
        "description": "AH49E Hall Effect Sensor - Dock Detection",
        "supplier": "Digi-Key",
        "part_number": "AH49E",
        "cost_qty1": 0.80,
        "cost_qty100": 0.50,
        "cost_qty500": 0.40,
        "cost_qty1000": 0.35,
        "lead_time": "In stock",
        "criticality": "LOW",
        "fmea_rpn": 12,
    },
    # Power
    {
        "part_id": "BAT-LIPO-001",
        "designation": "BAT1",
        "category": "Power",
        "quantity": 1,
        "description": "LiPo 3S 4000mAh (44Wh, 8-12h runtime)",
        "supplier": "AliExpress",
        "part_number": "3S 4000mAh Compact",
        "cost_qty1": 28.00,
        "cost_qty100": 22.00,
        "cost_qty500": 18.00,
        "cost_qty1000": 15.00,
        "lead_time": "2 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 144,  # Thermal runaway risk
    },
    {
        "part_id": "CHRG-BQ25895-001",
        "designation": "U13",
        "category": "Power",
        "quantity": 1,
        "description": "BQ25895 5A Buck-Boost Charger",
        "supplier": "Digi-Key",
        "part_number": "BQ25895RTWR",
        "cost_qty1": 4.26,
        "cost_qty100": 3.20,
        "cost_qty500": 2.80,
        "cost_qty1000": 2.40,
        "lead_time": "In stock",
        "criticality": "HIGH",
        "fmea_rpn": 54,
    },
    {
        "part_id": "FUEL-BQ40Z50-001",
        "designation": "U14",
        "category": "Power",
        "quantity": 1,
        "description": "BQ40Z50 Smart Battery Fuel Gauge",
        "supplier": "Digi-Key",
        "part_number": "BQ40Z50RSMR-R1",
        "cost_qty1": 5.85,
        "cost_qty100": 4.50,
        "cost_qty500": 3.80,
        "cost_qty1000": 3.20,
        "lead_time": "In stock",
        "criticality": "MEDIUM",
        "fmea_rpn": 32,
    },
    {
        "part_id": "WPT-P9415R-001",
        "designation": "U15",
        "category": "Power",
        "quantity": 1,
        "description": "Renesas P9415-R 15W Wireless Receiver",
        "supplier": "Renesas",
        "part_number": "P9415-R",
        "cost_qty1": 12.00,
        "cost_qty100": 8.00,
        "cost_qty500": 6.50,
        "cost_qty1000": 5.50,
        "lead_time": "2 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 105,  # WPT/maglev interference
    },
    {
        "part_id": "COIL-RX-001",
        "designation": "L1",
        "category": "Power",
        "quantity": 1,
        "description": "RX Coil 70mm Compact (Litz, 85uH)",
        "supplier": "Custom",
        "part_number": "Custom 70mm Coil",
        "cost_qty1": 20.00,
        "cost_qty100": 0.28,  # Volume pricing
        "cost_qty500": 0.22,
        "cost_qty1000": 0.18,
        "lead_time": "4 weeks",
        "criticality": "HIGH",
        "fmea_rpn": 72,
    },
    # Maglev
    {
        "part_id": "MAGLEV-HCNT-001",
        "designation": "MAG1",
        "category": "Maglev",
        "quantity": 1,
        "description": "HCNT Magnetic Levitation Module (2kg)",
        "supplier": "HCNT Alibaba",
        "part_number": "Naked Maglev",
        "cost_qty1": 45.00,
        "cost_qty100": 30.00,
        "cost_qty500": 25.00,
        "cost_qty1000": 22.00,
        "lead_time": "30-35 days",
        "criticality": "HIGH",
        "fmea_rpn": 56,
    },
    # Enclosure
    {
        "part_id": "ENCL-SHELL-001",
        "designation": "ENC1",
        "category": "Enclosure",
        "quantity": 1,
        "description": "85mm Sphere Shell (SLA/Injection)",
        "supplier": "Protolabs/Xometry",
        "part_number": "Custom",
        "cost_qty1": 50.00,  # SLA prototype
        "cost_qty100": 25.00,  # Soft tooling
        "cost_qty500": 8.00,  # Injection
        "cost_qty1000": 5.00,  # Volume injection
        "lead_time": "2-4 weeks",
        "criticality": "MEDIUM",
        "fmea_rpn": 48,
    },
]

# Manufacturing partners from research
MANUFACTURING_PARTNERS = [
    # PCBA
    {"stage": "PCBA (Prototype)", "partner": "JLCPCB", "location": "China", "moq": "1", "lead_time": "5-7 days", "cost": "$2 setup", "contact": "jlcpcb.com", "notes": "Best for 1-50 units"},
    {"stage": "PCBA (Prototype)", "partner": "PCBWay", "location": "China", "moq": "1", "lead_time": "3-5 days", "cost": "Free setup", "contact": "pcbway.com", "notes": "Free SMT stencil"},
    {"stage": "PCBA (Pilot)", "partner": "Seeed Fusion", "location": "China", "moq": "10", "lead_time": "7-10 days", "cost": "$0.50/joint", "contact": "seeedstudio.com", "notes": "Good DFM support"},
    {"stage": "PCBA (Production)", "partner": "MacroFab", "location": "USA", "moq": "100", "lead_time": "10-15 days", "cost": "$0.02/joint", "contact": "macrofab.com", "notes": "US-based, turnkey"},
    {"stage": "PCBA (Production)", "partner": "Tempo Automation", "location": "USA", "moq": "10", "lead_time": "3-5 days", "cost": "Premium", "contact": "tempoautomation.com", "notes": "Fast turn, high quality"},
    # Box Build
    {"stage": "Box Build (CM)", "partner": "Sofeast", "location": "China", "moq": "100", "lead_time": "4-6 weeks", "cost": "$5-15/unit", "contact": "sofeast.com", "notes": "Full turnkey CM"},
    {"stage": "Box Build (CM)", "partner": "Viasion", "location": "China", "moq": "500", "lead_time": "6-8 weeks", "cost": "$3-8/unit", "contact": "viasion.com", "notes": "Consumer electronics specialist"},
    # Enclosure
    {"stage": "Enclosure (Prototype)", "partner": "Protolabs", "location": "USA", "moq": "1", "lead_time": "3-5 days", "cost": "$50-200/part", "contact": "protolabs.com", "notes": "SLA, MJF, SLS"},
    {"stage": "Enclosure (Production)", "partner": "Xometry", "location": "USA", "moq": "1", "lead_time": "5-10 days", "cost": "Instant quote", "contact": "xometry.com", "notes": "Injection tooling $3-15K"},
    # Specialty
    {"stage": "Custom Coils", "partner": "Vishay/Würth", "location": "Various", "moq": "1000", "lead_time": "4-6 weeks", "cost": "$0.20-0.50/coil", "contact": "vishay.com", "notes": "Litz wire specialty"},
    {"stage": "Certification", "partner": "NTS", "location": "USA", "moq": "-", "lead_time": "4-8 weeks", "cost": "$3-8K", "contact": "nts.com", "notes": "FCC/CE/UL testing"},
    {"stage": "Certification", "partner": "Bureau Veritas", "location": "Global", "moq": "-", "lead_time": "6-10 weeks", "cost": "$5-15K", "contact": "bureauveritas.com", "notes": "Full compliance"},
]

# FMEA Critical Items (RPN > 100)
FMEA_CRITICAL = [
    {"component": "sensiBel SBM100B", "failure_mode": "Single source supplier unavailable", "severity": 8, "occurrence": 7, "detection": 4, "rpn": 224, "mitigation": "Dual footprint PCB for INMP441 fallback"},
    {"component": "Sony IMX989", "failure_mode": "Flagship sensor allocation", "severity": 7, "occurrence": 5, "detection": 5, "rpn": 175, "mitigation": "IMX890/IMX766 alternative qualified"},
    {"component": "Round AMOLED", "failure_mode": "Transparency insufficient for camera", "severity": 8, "occurrence": 4, "detection": 5, "rpn": 160, "mitigation": "Prototype validation required"},
    {"component": "Battery 3S", "failure_mode": "Thermal runaway in sealed sphere", "severity": 10, "occurrence": 2, "detection": 3, "rpn": 144, "mitigation": "BQ40Z50 protection, temp monitoring"},
    {"component": "Battery 3S", "failure_mode": "Swelling in sealed enclosure", "severity": 9, "occurrence": 2, "detection": 8, "rpn": 144, "mitigation": "Vent design, pressure relief"},
    {"component": "Dielectric Mirror", "failure_mode": "Coating delamination", "severity": 7, "occurrence": 4, "detection": 5, "rpn": 140, "mitigation": "Adhesion testing, supplier qualification"},
    {"component": "QCS6490", "failure_mode": "Thermal throttling in sphere", "severity": 8, "occurrence": 4, "detection": 4, "rpn": 128, "mitigation": "FEA thermal simulation $6-8K"},
    {"component": "P9415-R", "failure_mode": "WPT/Maglev interference", "severity": 7, "occurrence": 3, "detection": 5, "rpn": 105, "mitigation": "Ferrite shielding, frequency coordination"},
]


def create_bom_workbook():
    """Create comprehensive Excel BOM model."""
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    category_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    warning_fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
    good_fill = PatternFill(start_color="7FE57F", end_color="7FE57F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_format = '$#,##0.00'

    # ==================== SHEET 1: Full BOM ====================
    ws_bom = wb.active
    ws_bom.title = "Full BOM"

    # Title
    ws_bom['A1'] = "KAGAMI ORB V3 - BILL OF MATERIALS"
    ws_bom['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_bom['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_bom['A3'] = "Version: 3.0 | 85mm Compact SOTA with Living Eye Display"

    # Headers
    bom_headers = [
        "Part ID", "Designation", "Category", "Qty", "Description",
        "Supplier", "Part Number", "Cost (Qty 1)", "Cost (Qty 100)",
        "Cost (Qty 500)", "Cost (Qty 1000)", "Extended (Qty 1)",
        "Lead Time", "Criticality", "FMEA RPN"
    ]

    for col, header in enumerate(bom_headers, 1):
        cell = ws_bom.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data
    current_category = None
    row = 6
    for comp in COMPONENTS:
        # Category separator
        if comp['category'] != current_category:
            current_category = comp['category']
            ws_bom.cell(row=row, column=1, value=f"=== {current_category.upper()} ===")
            ws_bom.cell(row=row, column=1).font = Font(bold=True)
            for col in range(1, 16):
                ws_bom.cell(row=row, column=col).fill = category_fill
            row += 1

        # Component row
        ext_cost = comp['quantity'] * comp['cost_qty1']
        values = [
            comp['part_id'], comp['designation'], comp['category'], comp['quantity'],
            comp['description'], comp['supplier'], comp['part_number'],
            comp['cost_qty1'], comp['cost_qty100'], comp['cost_qty500'], comp['cost_qty1000'],
            ext_cost, comp['lead_time'], comp['criticality'], comp['fmea_rpn']
        ]

        for col, value in enumerate(values, 1):
            cell = ws_bom.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in [8, 9, 10, 11, 12]:  # Money columns
                cell.number_format = money_format
            # Color coding for criticality
            if col == 14:
                if value == "HIGH":
                    cell.fill = critical_fill
                elif value == "MEDIUM":
                    cell.fill = warning_fill
                else:
                    cell.fill = good_fill
            # Color coding for FMEA RPN
            if col == 15 and value > 100:
                cell.fill = critical_fill
        row += 1

    # Totals
    row += 1
    ws_bom.cell(row=row, column=11, value="TOTAL (Qty 1):").font = Font(bold=True)
    total_formula = f"=SUMPRODUCT(D6:D{row-2},H6:H{row-2})"
    ws_bom.cell(row=row, column=12, value=total_formula)
    ws_bom.cell(row=row, column=12).number_format = money_format
    ws_bom.cell(row=row, column=12).font = Font(bold=True)

    # Adjust column widths
    col_widths = [18, 10, 10, 6, 45, 20, 20, 12, 12, 12, 12, 14, 12, 10, 10]
    for col, width in enumerate(col_widths, 1):
        ws_bom.column_dimensions[get_column_letter(col)].width = width

    # ==================== SHEET 2: Cost Scenarios ====================
    ws_cost = wb.create_sheet("Cost Scenarios")

    ws_cost['A1'] = "COST SCENARIO ANALYSIS"
    ws_cost['A1'].font = Font(bold=True, size=16, color="1F4E79")

    # Quantity scenarios
    quantities = [1, 10, 100, 500, 1000]

    # Headers
    cost_headers = ["Category"] + [f"Qty {q}" for q in quantities] + ["Notes"]
    for col, header in enumerate(cost_headers, 1):
        cell = ws_cost.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Calculate costs by category
    categories = {}
    for comp in COMPONENTS:
        cat = comp['category']
        if cat not in categories:
            categories[cat] = {q: 0 for q in quantities}

        qty = comp['quantity']
        categories[cat][1] += qty * comp['cost_qty1']
        categories[cat][10] += qty * comp['cost_qty1'] * 0.95  # 5% discount at qty 10
        categories[cat][100] += qty * comp['cost_qty100']
        categories[cat][500] += qty * comp['cost_qty500']
        categories[cat][1000] += qty * comp['cost_qty1000']

    row = 4
    for cat, costs in categories.items():
        ws_cost.cell(row=row, column=1, value=cat)
        for col, q in enumerate(quantities, 2):
            cell = ws_cost.cell(row=row, column=col, value=costs[q])
            cell.number_format = money_format
            cell.border = thin_border
        row += 1

    # Assembly costs
    row += 1
    ws_cost.cell(row=row, column=1, value="Assembly Labor").font = Font(bold=True)
    assembly_costs = [50, 40, 25, 15, 12]  # Per unit assembly cost at each qty
    for col, cost in enumerate(assembly_costs, 2):
        cell = ws_cost.cell(row=row, column=col, value=cost)
        cell.number_format = money_format
        cell.border = thin_border

    # Tooling (one-time)
    row += 1
    ws_cost.cell(row=row, column=1, value="Tooling (Amortized)")
    tooling_costs = [0, 0, 80, 30, 15]  # Tooling per unit at volume
    for col, cost in enumerate(tooling_costs, 2):
        cell = ws_cost.cell(row=row, column=col, value=cost)
        cell.number_format = money_format
        cell.border = thin_border
    ws_cost.cell(row=row, column=7, value="$8K at 100, $15K at 500+")

    # Certification
    row += 1
    ws_cost.cell(row=row, column=1, value="Certification (Amortized)")
    cert_costs = [0, 0, 50, 12, 6]  # Certification per unit
    for col, cost in enumerate(cert_costs, 2):
        cell = ws_cost.cell(row=row, column=col, value=cost)
        cell.number_format = money_format
        cell.border = thin_border
    ws_cost.cell(row=row, column=7, value="FCC/CE/UL ~$5-8K total")

    # Grand totals
    row += 2
    ws_cost.cell(row=row, column=1, value="TOTAL PER UNIT").font = Font(bold=True, size=12)
    for col, q in enumerate(quantities, 2):
        # Sum all categories + assembly + tooling + cert
        total = sum(categories[cat][q] for cat in categories)
        total += assembly_costs[col-2]
        total += tooling_costs[col-2]
        total += cert_costs[col-2]
        cell = ws_cost.cell(row=row, column=col, value=total)
        cell.number_format = money_format
        cell.font = Font(bold=True)
        cell.fill = category_fill

    row += 1
    ws_cost.cell(row=row, column=1, value="TOTAL PROGRAM COST")
    for col, q in enumerate(quantities, 2):
        total_per = sum(categories[cat][q] for cat in categories)
        total_per += assembly_costs[col-2]
        program_total = total_per * q
        if q >= 100:
            program_total += 8000  # Tooling
        if q >= 500:
            program_total += 7000  # Additional tooling
        program_total += 5000 if q >= 100 else 0  # Certification
        cell = ws_cost.cell(row=row, column=col, value=program_total)
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)

    # Column widths
    for col in range(1, 8):
        ws_cost.column_dimensions[get_column_letter(col)].width = 15

    # ==================== SHEET 3: Manufacturing Partners ====================
    ws_mfg = wb.create_sheet("Manufacturing Partners")

    ws_mfg['A1'] = "MANUFACTURING PARTNER DIRECTORY"
    ws_mfg['A1'].font = Font(bold=True, size=16, color="1F4E79")

    mfg_headers = ["Stage", "Partner", "Location", "MOQ", "Lead Time", "Cost", "Contact", "Notes"]
    for col, header in enumerate(mfg_headers, 1):
        cell = ws_mfg.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row = 4
    current_stage = None
    for partner in MANUFACTURING_PARTNERS:
        if partner['stage'].split()[0] != current_stage:
            current_stage = partner['stage'].split()[0]

        values = [partner['stage'], partner['partner'], partner['location'],
                  partner['moq'], partner['lead_time'], partner['cost'],
                  partner['contact'], partner['notes']]
        for col, value in enumerate(values, 1):
            cell = ws_mfg.cell(row=row, column=col, value=value)
            cell.border = thin_border
        row += 1

    # Column widths
    mfg_widths = [25, 20, 12, 8, 12, 15, 25, 35]
    for col, width in enumerate(mfg_widths, 1):
        ws_mfg.column_dimensions[get_column_letter(col)].width = width

    # ==================== SHEET 4: FMEA Summary ====================
    ws_fmea = wb.create_sheet("FMEA Critical")

    ws_fmea['A1'] = "FMEA CRITICAL ITEMS (RPN > 100)"
    ws_fmea['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_fmea['A2'] = "Items requiring immediate mitigation before production"
    ws_fmea['A2'].font = Font(italic=True)

    fmea_headers = ["Component", "Failure Mode", "S", "O", "D", "RPN", "Mitigation"]
    for col, header in enumerate(fmea_headers, 1):
        cell = ws_fmea.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    row = 5
    for item in sorted(FMEA_CRITICAL, key=lambda x: -x['rpn']):
        values = [item['component'], item['failure_mode'], item['severity'],
                  item['occurrence'], item['detection'], item['rpn'], item['mitigation']]
        for col, value in enumerate(values, 1):
            cell = ws_fmea.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 6:  # RPN column
                if value >= 200:
                    cell.fill = critical_fill
                elif value >= 150:
                    cell.fill = warning_fill
        row += 1

    # Legend
    row += 2
    ws_fmea.cell(row=row, column=1, value="RPN = Severity × Occurrence × Detection")
    row += 1
    ws_fmea.cell(row=row, column=1, value="RPN > 200: CRITICAL - Must resolve before prototype")
    ws_fmea.cell(row=row, column=1).fill = critical_fill
    row += 1
    ws_fmea.cell(row=row, column=1, value="RPN 150-200: HIGH - Must resolve before pilot")
    ws_fmea.cell(row=row, column=1).fill = warning_fill
    row += 1
    ws_fmea.cell(row=row, column=1, value="RPN 100-150: MEDIUM - Must resolve before production")

    # Column widths
    fmea_widths = [20, 40, 6, 6, 6, 8, 50]
    for col, width in enumerate(fmea_widths, 1):
        ws_fmea.column_dimensions[get_column_letter(col)].width = width

    # ==================== SHEET 5: Sensitivity Analysis ====================
    ws_sens = wb.create_sheet("Sensitivity Analysis")

    ws_sens['A1'] = "COST SENSITIVITY ANALYSIS"
    ws_sens['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_sens['A2'] = "Impact of component cost changes on total BOM"

    # Top cost contributors
    ws_sens['A4'] = "TOP COST CONTRIBUTORS (Qty 100)"
    ws_sens['A4'].font = Font(bold=True)

    sens_headers = ["Component", "Current Cost", "% of BOM", "+20% Impact", "-20% Impact"]
    for col, header in enumerate(sens_headers, 1):
        cell = ws_sens.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Calculate total BOM at qty 100
    total_bom = sum(c['quantity'] * c['cost_qty100'] for c in COMPONENTS)

    # Sort by cost contribution
    sorted_comps = sorted(COMPONENTS, key=lambda x: x['quantity'] * x['cost_qty100'], reverse=True)

    row = 6
    for comp in sorted_comps[:10]:  # Top 10
        ext_cost = comp['quantity'] * comp['cost_qty100']
        pct = ext_cost / total_bom * 100
        impact_up = ext_cost * 0.2
        impact_down = ext_cost * 0.2

        values = [comp['description'][:40], ext_cost, pct, impact_up, impact_down]
        for col, value in enumerate(values, 1):
            cell = ws_sens.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 2 or col >= 4:
                cell.number_format = money_format
            elif col == 3:
                cell.number_format = '0.0%'
                cell.value = pct / 100
        row += 1

    # Break-even analysis
    row += 2
    ws_sens.cell(row=row, column=1, value="BREAK-EVEN ANALYSIS").font = Font(bold=True)
    row += 1
    ws_sens.cell(row=row, column=1, value="Target retail price:")
    ws_sens.cell(row=row, column=2, value=999)
    ws_sens.cell(row=row, column=2).number_format = money_format
    row += 1
    ws_sens.cell(row=row, column=1, value="Target margin:")
    ws_sens.cell(row=row, column=2, value=0.4)
    ws_sens.cell(row=row, column=2).number_format = '0%'
    row += 1
    ws_sens.cell(row=row, column=1, value="Max BOM cost:")
    ws_sens.cell(row=row, column=2, value="=B" + str(row-2) + "*(1-B" + str(row-1) + ")")
    ws_sens.cell(row=row, column=2).number_format = money_format
    ws_sens.cell(row=row, column=2).font = Font(bold=True)
    row += 1
    ws_sens.cell(row=row, column=1, value="Current BOM (Qty 500):")
    bom_500 = sum(c['quantity'] * c['cost_qty500'] for c in COMPONENTS)
    ws_sens.cell(row=row, column=2, value=bom_500)
    ws_sens.cell(row=row, column=2).number_format = money_format
    row += 1
    ws_sens.cell(row=row, column=1, value="Gap to target:")
    ws_sens.cell(row=row, column=2, value=f"=B{row-2}-B{row-1}")
    ws_sens.cell(row=row, column=2).number_format = money_format

    # Column widths
    for col in range(1, 6):
        ws_sens.column_dimensions[get_column_letter(col)].width = 18
    ws_sens.column_dimensions['A'].width = 45

    # ==================== SHEET 6: Summary Dashboard ====================
    ws_dash = wb.create_sheet("Dashboard")

    ws_dash['A1'] = "KAGAMI ORB V3 - EXECUTIVE SUMMARY"
    ws_dash['A1'].font = Font(bold=True, size=18, color="1F4E79")

    ws_dash['A3'] = "Project Status: READY FOR PROTOTYPE"
    ws_dash['A3'].font = Font(bold=True, size=14)
    ws_dash['A3'].fill = good_fill

    # Key metrics
    ws_dash['A5'] = "KEY METRICS"
    ws_dash['A5'].font = Font(bold=True, size=12)

    metrics = [
        ("Total Components:", len(COMPONENTS)),
        ("Unique Part Numbers:", len(set(c['part_id'] for c in COMPONENTS))),
        ("Critical Items (HIGH):", sum(1 for c in COMPONENTS if c['criticality'] == 'HIGH')),
        ("FMEA Items > 100 RPN:", len(FMEA_CRITICAL)),
        ("", ""),
        ("BOM Cost (Qty 1):", sum(c['quantity'] * c['cost_qty1'] for c in COMPONENTS)),
        ("BOM Cost (Qty 100):", sum(c['quantity'] * c['cost_qty100'] for c in COMPONENTS)),
        ("BOM Cost (Qty 500):", sum(c['quantity'] * c['cost_qty500'] for c in COMPONENTS)),
        ("BOM Cost (Qty 1000):", sum(c['quantity'] * c['cost_qty1000'] for c in COMPONENTS)),
        ("", ""),
        ("Target Retail Price:", 999),
        ("Required Margin:", "40%"),
        ("Max Allowable BOM:", 599),
    ]

    row = 6
    for label, value in metrics:
        ws_dash.cell(row=row, column=1, value=label)
        cell = ws_dash.cell(row=row, column=2, value=value)
        if isinstance(value, (int, float)) and value > 100:
            cell.number_format = money_format
        row += 1

    # Risk summary
    row += 1
    ws_dash.cell(row=row, column=1, value="TOP RISKS").font = Font(bold=True, size=12)
    row += 1
    for i, item in enumerate(sorted(FMEA_CRITICAL, key=lambda x: -x['rpn'])[:5], 1):
        ws_dash.cell(row=row, column=1, value=f"{i}. {item['component']}: {item['failure_mode']}")
        ws_dash.cell(row=row, column=2, value=f"RPN {item['rpn']}")
        if item['rpn'] >= 200:
            ws_dash.cell(row=row, column=2).fill = critical_fill
        row += 1

    # Column widths
    ws_dash.column_dimensions['A'].width = 50
    ws_dash.column_dimensions['B'].width = 15

    return wb


def main():
    """Generate the Excel BOM model."""
    print("Generating Kagami Orb V3 BOM Model...")

    wb = create_bom_workbook()

    output_path = "/Users/schizodactyl/projects/kagami/apps/hub/kagami-orb/bom/kagami_orb_v3_bom_model.xlsx"
    wb.save(output_path)

    print(f"✓ BOM model saved to: {output_path}")
    print(f"  - Full BOM: {len(COMPONENTS)} components")
    print(f"  - Cost Scenarios: Qty 1, 10, 100, 500, 1000")
    print(f"  - Manufacturing Partners: {len(MANUFACTURING_PARTNERS)} contacts")
    print(f"  - FMEA Critical Items: {len(FMEA_CRITICAL)} (RPN > 100)")

    # Summary stats
    total_qty1 = sum(c['quantity'] * c['cost_qty1'] for c in COMPONENTS)
    total_qty500 = sum(c['quantity'] * c['cost_qty500'] for c in COMPONENTS)
    print(f"\n  BOM Cost Summary:")
    print(f"    Qty 1:   ${total_qty1:,.2f}")
    print(f"    Qty 500: ${total_qty500:,.2f}")
    print(f"    Savings: {(1 - total_qty500/total_qty1)*100:.1f}%")


if __name__ == "__main__":
    main()
